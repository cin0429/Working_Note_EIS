from __future__ import annotations

import argparse
import re
from copy import copy
from collections import defaultdict
from difflib import get_close_matches
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Special categories are summarized into another sheet, not written to EIS.
SPECIAL_CATEGORIES_NORM = {"kore", "other", "others", "rfi rfq"}

HEADER_ROW = 1
EIS_ITEM_COL = 1

BRACKET_ONLY_RE = re.compile(r"^\s*\[[^\]]+\]\s*$")
TRAILING_QTY_RE = re.compile(r"\s*(?:-\s*)?x\s*(\d+(?:\.\d+)?)\s*$", re.IGNORECASE)
STATUS_RE = re.compile(r"\[[^\]]+\]")


def norm(text: object) -> str:
    s = str(text or "").strip().lower()
    s = STATUS_RE.sub(" ", s)
    s = s.replace("/", " ").replace("_", " ")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def display_qty(qty: float):
    return int(qty) if float(qty).is_integer() else qty


def display_category(project: str) -> str:
    n = norm(project)
    if n in {"other", "others"}:
        return "Other"
    if n == "rfi rfq":
        return "RFI_RFQ"
    if n == "kore":
        return "Kore"
    return project.strip()


def find_work_item_col(ws: Worksheet) -> int:
    for cell in ws[HEADER_ROW]:
        if cell.value and "工作項目" in str(cell.value):
            return cell.column
    return 5


def has_summary_header(ws: Worksheet) -> bool:
    """Return True if this worksheet looks like the main EIS table."""
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(HEADER_ROW, c).value) == "summary":
            return True
    return False


def get_eis_sheet(wb, sheet_name: Optional[str] = None) -> Worksheet:
    """
    Pick the main EIS worksheet.

    Important: when users add Project_Alias / Item_Alias and save Excel while that
    sheet is selected, openpyxl may treat the alias sheet as wb.active. Therefore
    we must not blindly use wb.active. If --eis-sheet is not provided, choose the
    first worksheet whose first row contains a Summary header.
    """
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Cannot find EIS sheet: {sheet_name}")
        ws = wb[sheet_name]
        if not has_summary_header(ws):
            raise ValueError(f"Sheet '{sheet_name}' does not contain a Summary header in row {HEADER_ROW}.")
        return ws

    # Prefer the active sheet only if it is truly the EIS table.
    if has_summary_header(wb.active):
        return wb.active

    # Otherwise search all non-helper sheets.
    helper_sheets = {"project_alias", "item_alias"}
    for ws in wb.worksheets:
        if norm(ws.title) in helper_sheets:
            continue
        if has_summary_header(ws):
            return ws

    # Last chance: search every sheet, including helper sheets, and report clearly.
    for ws in wb.worksheets:
        if has_summary_header(ws):
            return ws

    available = ", ".join(wb.sheetnames)
    raise ValueError(
        "Cannot find the main EIS sheet. No worksheet has a 'Summary' header in row "
        f"{HEADER_ROW}. Available sheets: {available}"
    )


def build_eis_maps(ws: Worksheet) -> Tuple[Dict[str, int], Dict[str, int]]:
    item_row_by_norm: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, EIS_ITEM_COL).value
        if val:
            item_row_by_norm[norm(val)] = r

    project_col_by_norm: Dict[str, int] = {}
    for c in range(3, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, c).value
        if val and norm(val) != "summary":
            project_col_by_norm[norm(val)] = c
    return item_row_by_norm, project_col_by_norm




def find_summary_col(ws: Worksheet) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(HEADER_ROW, c).value) == "summary":
            return c
    return None


def find_project_template_col(ws: Worksheet) -> int:
    """Use the last normal project column before Summary as the style template."""
    summary_col = find_summary_col(ws) or (ws.max_column + 1)
    for c in range(summary_col - 1, 2, -1):
        if ws.cell(HEADER_ROW, c).value:
            return c
    return max(3, summary_col - 1)


def copy_column_format(ws: Worksheet, src_col: int, dst_col: int) -> None:
    """Copy visual formatting only. Values/formulas are not copied."""
    src_letter = ws.cell(1, src_col).column_letter
    dst_letter = ws.cell(1, dst_col).column_letter
    ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
    ws.column_dimensions[dst_letter].hidden = ws.column_dimensions[src_letter].hidden
    for r in range(1, ws.max_row + 1):
        src = ws.cell(r, src_col)
        dst = ws.cell(r, dst_col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def add_project_column_preserve_structure(ws: Worksheet, project_name: str, project_col_by_norm: Dict[str, int]) -> int:
    """
    Add a new project column immediately before the Summary column.

    This keeps the EIS structure as:
        Item columns + Project columns + Summary

    Existing project columns stay in the same order, Summary is shifted one column to the right,
    and the inserted project column copies the style from the previous project column.
    """
    summary_col = find_summary_col(ws)
    if not summary_col:
        raise ValueError('Cannot find Summary column in EIS header row. Please confirm the header is named Summary.')

    # Insert a new column exactly before Summary. Existing Summary moves one column right.
    dst_col = summary_col
    template_col = find_project_template_col(ws)
    ws.insert_cols(dst_col, 1)

    # After insertion, copy format from the project column immediately to the left when possible.
    # If there is no project column on the left, fall back to the original template column shifted by insertion.
    src_col = max(3, dst_col - 1)
    copy_column_format(ws, src_col, dst_col)

    # Clear any copied values/formulas in the new project column and keep only formatting.
    for r in range(1, ws.max_row + 1):
        ws.cell(r, dst_col).value = None

    ws.cell(HEADER_ROW, dst_col).value = project_name.strip()

    # Existing columns at and after dst_col shifted right, so refresh project map instead of patching indexes.
    _, refreshed_project_map = build_eis_maps(ws)
    project_col_by_norm.clear()
    project_col_by_norm.update(refreshed_project_map)
    return dst_col


def update_summary_formulas_for_new_projects(ws: Worksheet, new_project_cols: List[int]) -> None:
    """
    Rebuild Summary formulas after inserting new project columns before Summary.

    Because openpyxl does not reliably update formulas for inserted columns, this function
    explicitly makes Summary sum every project column between C and the column before Summary.
    It only runs when new project columns were created.
    """
    if not new_project_cols:
        return
    summary_col = find_summary_col(ws)
    if not summary_col:
        return

    first_project_col = 3
    last_project_col = summary_col - 1
    if last_project_col < first_project_col:
        return

    for r in range(2, ws.max_row + 1):
        # Only update real EIS item rows. This avoids accidentally writing formulas into blank footer rows.
        if not ws.cell(r, EIS_ITEM_COL).value:
            continue
        first_cell = ws.cell(r, first_project_col).coordinate
        last_cell = ws.cell(r, last_project_col).coordinate
        ws.cell(r, summary_col).value = f"=SUM({first_cell}:{last_cell})"


def read_alias_sheet(wb, sheet_name: str, key_header: str, value_header: str) -> Dict[str, str]:
    """Read alias sheet if it exists. It does not create or modify the EIS page."""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    headers = {norm(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    key_col = headers.get(norm(key_header))
    value_col = headers.get(norm(value_header))
    if not key_col or not value_col:
        return {}

    alias_map: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        alias = ws.cell(r, key_col).value
        target = ws.cell(r, value_col).value
        if alias and target:
            alias_map[norm(alias)] = str(target).strip()
    return alias_map


def resolve_task_row(task_text: str, item_row_by_norm: Dict[str, int], item_aliases: Dict[str, str]) -> Optional[int]:
    n = norm(task_text)
    if not n:
        return None

    alias_target = item_aliases.get(n)
    if alias_target:
        return item_row_by_norm.get(norm(alias_target))

    if n in item_row_by_norm:
        return item_row_by_norm[n]

    # Containment match: useful when employee only writes part of a longer EIS item.
    for eis_norm, row in item_row_by_norm.items():
        if eis_norm and (eis_norm in n or n in eis_norm):
            return row

    # Token subset match.
    n_tokens = set(n.split())
    for eis_norm, row in item_row_by_norm.items():
        eis_tokens = set(eis_norm.split())
        if not n_tokens or not eis_tokens:
            continue
        if len(n_tokens) >= 4 and n_tokens.issubset(eis_tokens):
            return row
        if len(eis_tokens) >= 4 and eis_tokens.issubset(n_tokens):
            return row

    # Conservative fuzzy fallback.
    matches = get_close_matches(n, list(item_row_by_norm.keys()), n=1, cutoff=0.88)
    if matches:
        return item_row_by_norm[matches[0]]
    return None


def parse_dash_line(line: str) -> Optional[Tuple[str, str, float]]:
    """Parse: Project - Task - xN or Project - Task. Missing quantity defaults to x1."""
    line = line.strip()
    if not line or BRACKET_ONLY_RE.match(line):
        return None
    if "-" not in line:
        return None

    project, task_part = re.split(r"\s*-\s*", line, maxsplit=1)
    project = project.strip()
    task_part = task_part.strip()
    if not project or not task_part:
        return None

    qty = 1.0
    qty_match = TRAILING_QTY_RE.search(task_part)
    if qty_match:
        qty = float(qty_match.group(1))
        task_part = TRAILING_QTY_RE.sub("", task_part).strip()

    task_part = STATUS_RE.sub("", task_part)
    task_part = re.sub(r"\s+", " ", task_part).strip(" -")
    if not task_part:
        return None
    return project, task_part, qty


def parse_work_note_cell(text: object) -> List[Tuple[str, str, float]]:
    if not text:
        return []
    results: List[Tuple[str, str, float]] = []
    for raw_line in str(text).splitlines():
        if raw_line and raw_line[0].isspace():
            continue
        parsed = parse_dash_line(raw_line)
        if parsed:
            results.append(parsed)
    return results


def read_working_notes(paths: Iterable[Path]) -> List[Tuple[str, str, float, str, str, int]]:
    records = []
    for path in paths:
        wb = load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            work_col = find_work_item_col(ws)
            for r in range(2, ws.max_row + 1):
                cell_text = ws.cell(r, work_col).value
                for project, task, qty in parse_work_note_cell(cell_text):
                    records.append((project, task, qty, path.name, ws.title, r))
    return records


def recreate_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    return wb.create_sheet(sheet_name)


def write_summary_sheets(wb, special_counts, unknown_projects, unknown_items, unmatched_detail, process_log):
    ws = recreate_sheet(wb, "Special_Category_Summary")
    ws.append(["Category", "Task", "Quantity"])
    for (cat, task), qty in sorted(special_counts.items(), key=lambda x: (x[0][0], x[0][1])):
        ws.append([cat, task, display_qty(qty)])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 75
    ws.column_dimensions["C"].width = 12

    ws2 = recreate_sheet(wb, "Unknown_Project_Summary")
    ws2.append(["Project", "Task", "Quantity"])
    for (project, task), qty in sorted(unknown_projects.items(), key=lambda x: (x[0][0], x[0][1])):
        ws2.append([project, task, display_qty(qty)])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 75
    ws2.column_dimensions["C"].width = 12

    ws3 = recreate_sheet(wb, "Unknown_Item_Summary")
    ws3.append(["Project", "Task", "Quantity"])
    for (project, task), qty in sorted(unknown_items.items(), key=lambda x: (x[0][0], x[0][1])):
        ws3.append([project, task, display_qty(qty)])
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 75
    ws3.column_dimensions["C"].width = 12

    ws4 = recreate_sheet(wb, "Unmatched_WorkingNote_Detail")
    ws4.append(["Project", "Task", "Quantity", "Source File", "Source Sheet", "Source Row", "Reason"])
    for row in unmatched_detail:
        ws4.append(row)
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 75
    ws4.column_dimensions["D"].width = 28
    ws4.column_dimensions["G"].width = 35

    ws5 = recreate_sheet(wb, "Process_Log")
    ws5.append(["Project", "Task", "Quantity", "Source File", "Source Sheet", "Source Row", "Status", "Matched EIS Row", "Matched EIS Project Column"])
    for row in process_log:
        ws5.append(row)
    ws5.column_dimensions["A"].width = 22
    ws5.column_dimensions["B"].width = 75
    ws5.column_dimensions["D"].width = 28
    ws5.column_dimensions["G"].width = 24


def calculate(eis_path: Path, working_note_paths: List[Path], output_path: Path, eis_sheet_name: Optional[str] = None) -> None:
    eis_wb = load_workbook(eis_path)
    eis_ws = get_eis_sheet(eis_wb, eis_sheet_name)
    item_row_by_norm, project_col_by_norm = build_eis_maps(eis_ws)

    # Optional alias sheets. These can exist in EIS.xlsx, but this script will not create/change the EIS page format.
    project_aliases = read_alias_sheet(eis_wb, "Project_Alias", "Alias", "Real_Project")
    item_aliases = read_alias_sheet(eis_wb, "Item_Alias", "Alias", "EIS_Item")

    records = read_working_notes(working_note_paths)
    special_counts = defaultdict(float)
    unknown_projects = defaultdict(float)
    unknown_items = defaultdict(float)
    unmatched_detail = []
    process_log = []
    new_project_cols: List[int] = []

    for project, task, qty, src_file, src_sheet, src_row in records:
        original_project = project.strip()
        project = project_aliases.get(norm(project), original_project)
        project_norm = norm(project)

        if project_norm in SPECIAL_CATEGORIES_NORM:
            category = display_category(project)
            special_counts[(category, task)] += qty
            process_log.append([category, task, display_qty(qty), src_file, src_sheet, src_row, "Special Category", "", ""])
            continue

        col = project_col_by_norm.get(project_norm)
        if not col:
            # New project is allowed. Insert it immediately before Summary so Summary stays last.
            # Existing EIS page structure/style is preserved as much as possible; only the new column
            # is added and formatted like the neighboring project column.
            col = add_project_column_preserve_structure(eis_ws, project.strip(), project_col_by_norm)
            new_project_cols.append(col)
            unknown_projects[(project.strip(), task)] += qty
            process_log.append([project.strip(), task, display_qty(qty), src_file, src_sheet, src_row, "New Project Column Added", "", col])

        row = resolve_task_row(task, item_row_by_norm, item_aliases)
        if not row:
            unknown_items[(project.strip(), task)] += qty
            unmatched_detail.append([project.strip(), task, display_qty(qty), src_file, src_sheet, src_row, "Cannot match EIS item in column A"])
            process_log.append([project.strip(), task, display_qty(qty), src_file, src_sheet, src_row, "Unknown Item", "", col])
            continue

        old = eis_ws.cell(row, col).value
        old_num = float(old) if isinstance(old, (int, float)) else 0.0
        new_value = old_num + qty
        # Only value is written into an existing EIS cell. No row/column insertion and no style/format changes.
        eis_ws.cell(row, col).value = display_qty(new_value)
        process_log.append([project.strip(), task, display_qty(qty), src_file, src_sheet, src_row, "Matched", row, col])

    update_summary_formulas_for_new_projects(eis_ws, new_project_cols)
    write_summary_sheets(eis_wb, special_counts, unknown_projects, unknown_items, unmatched_detail, process_log)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    eis_wb.save(output_path)
    print(f"Parsed records: {len(records)}")
    print(f"Special category groups: {len(special_counts)}")
    print(f"Unknown project groups: {len(unknown_projects)}")
    print(f"Unknown item groups: {len(unknown_items)}")
    print(f"Done: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Calculate monthly EIS quantity from WorkingNote Excel files. New project columns are inserted before Summary.")
    parser.add_argument("--eis", required=True, help="Path to EIS.xlsx template/monthly file")
    parser.add_argument("--working-notes", nargs="+", required=True, help="One or more WorkingNote.xlsx files")
    parser.add_argument("--output", required=True, help="Output xlsx path")
    parser.add_argument("--eis-sheet", default=None, help="Optional EIS sheet name. Default = active sheet")
    args = parser.parse_args()

    calculate(Path(args.eis), [Path(p) for p in args.working_notes], Path(args.output), args.eis_sheet)


if __name__ == "__main__":
    main()
