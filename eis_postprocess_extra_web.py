# -*- coding: utf-8 -*-
"""
eis_postprocess_extra_web.py

Purpose:
    Post-process reviewed EIS_calculated.xlsx and fill:
      1) EIS_Extra
      2) EIS_Web

Input:
    EIS_calculated.xlsx  (default)

Output:
    EIS_final.xlsx       (default)

Expected workbook sheets:
    - EIS main sheet, or any sheet whose row 1 contains "Summary" and row 20 contains "掛名人員"
    - EIS_Extra
    - EIS_Web
    - Project_Alias       optional, format: Alias | Real_Project | Project Code
    - Employee_Alias      optional, format: Alias | Real_Name

Rules:
    EIS_Extra:
      - Scan EIS row 20 for cells containing "Extra".
      - For each matched column:
          EIS_Extra column A = EIS row 1 project name
          EIS_Extra column B = Project Code from Project_Alias
          EIS_Extra column C = EIS row 19 建議申報EIS value

    EIS_Web:
      - For each row in EIS_Web, starting from row 2 by default:
          A column = employee name, e.g. Amber_Chen(陳圓圓_Company)
          D column = project name
      - Match employee against EIS row 20 掛名人員.
      - Match project against EIS row 1 project name.
      - If both match in same EIS column, write EIS row 19 value to EIS_Web column E.
      - If not found, write "0".

Usage:
    python eis_postprocess_extra_web.py
    python eis_postprocess_extra_web.py --input EIS_calculated.xlsx --output EIS_final.xlsx
    python eis_postprocess_extra_web.py --eis-sheet EIS
"""

from __future__ import annotations

import argparse
import re
from copy import copy
from pathlib import Path
from typing import Dict, Optional, Tuple, Any, List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_INPUT_FILE = "EIS_calculated.xlsx"
DEFAULT_OUTPUT_FILE = "EIS_final.xlsx"

EIS_EXTRA_SHEET = "EIS_Extra"
EIS_WEB_SHEET = "EIS_Web"
PROJECT_ALIAS_SHEET = "Project_Alias"
EMPLOYEE_ALIAS_SHEET = "Employee_Alias"

PROJECT_ROW = 1
EIS_VALUE_ROW = 19
ASSIGNEE_ROW = 20

# EIS_Web column settings
EIS_WEB_START_ROW = 2
EIS_WEB_EMPLOYEE_COL = 1  # A
EIS_WEB_PROJECT_COL = 4   # D
EIS_WEB_OUTPUT_COL = 5    # E

# EIS_Extra column settings
EIS_EXTRA_START_ROW = 2
EIS_EXTRA_PROJECT_COL = 1       # A
EIS_EXTRA_PROJECT_CODE_COL = 2  # B
EIS_EXTRA_VALUE_COL = 3         # C


# -----------------------------
# Text normalization helpers
# -----------------------------

def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_basic(value: Any) -> str:
    """Case-insensitive, whitespace-normalized key."""
    text = cell_text(value).lower()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def norm_loose(value: Any) -> str:
    """Looser key for matching project / employee names."""
    text = norm_basic(value)
    text = text.replace("_", " ")
    text = re.sub(r"[\[\]\(\)（）{}<>:;,'\"`~!@#$%^&*+=|\\/?]+", " ", text)
    text = re.sub(r"[-]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def text_contains_either(a: str, b: str) -> bool:
    """Return True if normalized a contains b or b contains a."""
    na = norm_loose(a)
    nb = norm_loose(b)
    if not na or not nb:
        return False
    return na in nb or nb in na


# -----------------------------
# Alias loading
# -----------------------------

def load_alias_sheet(wb: openpyxl.Workbook, sheet_name: str) -> Dict[str, str]:
    """
    Load a two-column alias sheet.
    Column A = Alias
    Column B = Real value
    Row 1 is header.
    """
    alias_map: Dict[str, str] = {}
    if sheet_name not in wb.sheetnames:
        return alias_map

    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        alias, real = row[0], row[1]
        if cell_text(alias) and cell_text(real):
            alias_map[norm_loose(alias)] = cell_text(real)
            alias_map[norm_basic(alias)] = cell_text(real)
    return alias_map



def load_project_alias_and_codes(wb: openpyxl.Workbook) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Load Project_Alias sheet.

    Expected columns:
      A = Alias
      B = Real_Project
      C = Project Code

    Rules:
      - Alias is used to normalize project names.
      - Project Code can be found by either Alias or Real_Project.
      - If C column is empty, the alias still works, but Project Code will be blank.
    """
    alias_map: Dict[str, str] = {}
    code_map: Dict[str, str] = {}

    if PROJECT_ALIAS_SHEET not in wb.sheetnames:
        return alias_map, code_map

    ws = wb[PROJECT_ALIAS_SHEET]
    for row in ws.iter_rows(min_row=2, values_only=True):
        alias = row[0] if len(row) >= 1 else None
        real_project = row[1] if len(row) >= 2 else None
        project_code = row[2] if len(row) >= 3 else None

        alias_text = cell_text(alias)
        real_text = cell_text(real_project)
        code_text = cell_text(project_code)

        if alias_text and real_text:
            alias_map[norm_loose(alias_text)] = real_text
            alias_map[norm_basic(alias_text)] = real_text

        # Allow Project Code lookup by both alias and official project name.
        if code_text:
            if alias_text:
                code_map[norm_loose(alias_text)] = code_text
                code_map[norm_basic(alias_text)] = code_text
            if real_text:
                code_map[norm_loose(real_text)] = code_text
                code_map[norm_basic(real_text)] = code_text

    return alias_map, code_map


def get_project_code(project_name: Any, project_alias_map: Dict[str, str], project_code_map: Dict[str, str]) -> str:
    """Return Project Code using alias-normalized project name."""
    raw = cell_text(project_name)
    if not raw:
        return ""

    # Try original project name first.
    code = project_code_map.get(norm_loose(raw), project_code_map.get(norm_basic(raw), ""))
    if code:
        return code

    # Then try canonical project name after applying Project_Alias.
    canonical = apply_alias(raw, project_alias_map)
    return project_code_map.get(norm_loose(canonical), project_code_map.get(norm_basic(canonical), ""))


def apply_alias(value: Any, alias_map: Dict[str, str]) -> str:
    raw = cell_text(value)
    if not raw:
        return ""
    return alias_map.get(norm_loose(raw), alias_map.get(norm_basic(raw), raw))


def normalize_employee_name(value: Any, employee_alias_map: Dict[str, str]) -> str:
    """
    Normalize EIS_Web employee name.
    Priority:
      1. Employee_Alias sheet exact/loose match
      2. Take text before '(' or '（'
      3. Take text before '_' from the remaining part

    Example:
      Amber_Chen(陳圓圓_Company) -> Amber
    """
    raw = cell_text(value)
    if not raw:
        return ""

    aliased = apply_alias(raw, employee_alias_map)
    if aliased != raw:
        return aliased

    base = re.split(r"[\(（]", raw, maxsplit=1)[0].strip()
    if "_" in base:
        return base.split("_", 1)[0].strip()
    return base


# -----------------------------
# Workbook / sheet detection
# -----------------------------

def find_eis_sheet(wb: openpyxl.Workbook, preferred_sheet: Optional[str] = None) -> Worksheet:
    if preferred_sheet:
        if preferred_sheet not in wb.sheetnames:
            raise ValueError(f'Cannot find EIS sheet "{preferred_sheet}" in workbook.')
        return wb[preferred_sheet]

    # Prefer sheet named EIS if it looks valid.
    if "EIS" in wb.sheetnames:
        ws = wb["EIS"]
        if row_contains(ws, PROJECT_ROW, "Summary") or row_contains(ws, ASSIGNEE_ROW, "掛名人員"):
            return ws

    # Auto-detect: row 1 has Summary and row 20 has 掛名人員.
    for ws in wb.worksheets:
        if ws.title in {EIS_EXTRA_SHEET, EIS_WEB_SHEET, PROJECT_ALIAS_SHEET, EMPLOYEE_ALIAS_SHEET}:
            continue
        if row_contains(ws, PROJECT_ROW, "Summary") and row_contains(ws, ASSIGNEE_ROW, "掛名人員"):
            return ws

    # Fallback: any sheet with Summary in row 1.
    for ws in wb.worksheets:
        if ws.title in {EIS_EXTRA_SHEET, EIS_WEB_SHEET, PROJECT_ALIAS_SHEET, EMPLOYEE_ALIAS_SHEET}:
            continue
        if row_contains(ws, PROJECT_ROW, "Summary"):
            return ws

    raise ValueError("Cannot auto-detect EIS sheet. Please use --eis-sheet <sheet_name>.")


def row_contains(ws: Worksheet, row: int, keyword: str) -> bool:
    key = norm_basic(keyword)
    for col in range(1, ws.max_column + 1):
        if key in norm_basic(ws.cell(row, col).value):
            return True
    return False


def ensure_sheet(wb: openpyxl.Workbook, sheet_name: str) -> Worksheet:
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)


# -----------------------------
# Style helpers
# -----------------------------

def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def clear_old_body(ws: Worksheet, start_row: int, columns: List[int]) -> None:
    for row in range(start_row, ws.max_row + 1):
        for col in columns:
            ws.cell(row, col).value = None


def write_with_style(ws: Worksheet, row: int, col: int, value: Any, style_from_row: Optional[int] = None) -> None:
    cell = ws.cell(row, col)
    cell.value = value
    if style_from_row and style_from_row >= 1:
        src = ws.cell(style_from_row, col)
        copy_cell_style(src, cell)


# -----------------------------
# Lookup building
# -----------------------------

def build_eis_column_records(
    eis_ws: Worksheet,
    eis_values_ws: Worksheet,
    project_alias_map: Dict[str, str],
) -> List[Dict[str, Any]]:
    """Build list of column records from EIS row 1 / row 19 / row 20."""
    records: List[Dict[str, Any]] = []
    for col in range(1, eis_ws.max_column + 1):
        project = cell_text(eis_ws.cell(PROJECT_ROW, col).value)
        assignee = cell_text(eis_ws.cell(ASSIGNEE_ROW, col).value)
        value = eis_values_ws.cell(EIS_VALUE_ROW, col).value

        if not project and not assignee:
            continue

        # Skip Summary as a Project candidate for EIS_Web, but keep data harmless.
        canonical_project = apply_alias(project, project_alias_map)
        records.append({
            "col": col,
            "project": project,
            "canonical_project": canonical_project,
            "project_key": norm_loose(canonical_project),
            "assignee": assignee,
            "assignee_key": norm_loose(assignee),
            "value": value,
        })
    return records


# -----------------------------
# EIS_Extra processing
# -----------------------------

def update_eis_extra(
    wb: openpyxl.Workbook,
    records: List[Dict[str, Any]],
    project_alias_map: Dict[str, str],
    project_code_map: Dict[str, str],
) -> int:
    ws = ensure_sheet(wb, EIS_EXTRA_SHEET)

    # Preserve header, clear previous generated data in A and C.
    clear_old_body(ws, EIS_EXTRA_START_ROW, [EIS_EXTRA_PROJECT_COL, EIS_EXTRA_PROJECT_CODE_COL, EIS_EXTRA_VALUE_COL])

    out_row = EIS_EXTRA_START_ROW
    for rec in records:
        project = rec["project"]
        assignee = rec["assignee"]
        if not project:
            continue
        if "extra" in norm_basic(assignee):
            project_code = get_project_code(project, project_alias_map, project_code_map)
            write_with_style(ws, out_row, EIS_EXTRA_PROJECT_COL, project, EIS_EXTRA_START_ROW)
            write_with_style(ws, out_row, EIS_EXTRA_PROJECT_CODE_COL, project_code, EIS_EXTRA_START_ROW)
            write_with_style(ws, out_row, EIS_EXTRA_VALUE_COL, rec["value"], EIS_EXTRA_START_ROW)
            out_row += 1

    return out_row - EIS_EXTRA_START_ROW


# -----------------------------
# EIS_Web processing
# -----------------------------

def project_matches(request_project: str, rec_project: str, project_alias_map: Dict[str, str]) -> bool:
    req = apply_alias(request_project, project_alias_map)
    rec = apply_alias(rec_project, project_alias_map)
    if not req or not rec:
        return False
    if norm_loose(req) == norm_loose(rec):
        return True
    # Allow partial match similar to old Excel SEARCH logic, but after alias normalization.
    return text_contains_either(req, rec)


def employee_matches(request_employee: str, eis_assignee: str, employee_alias_map: Dict[str, str]) -> bool:
    req = normalize_employee_name(request_employee, employee_alias_map)
    eis = apply_alias(eis_assignee, employee_alias_map)
    if not req or not eis:
        return False
    if norm_loose(req) == norm_loose(eis):
        return True
    return text_contains_either(req, eis)


def find_eis_web_value(
    employee: str,
    project: str,
    records: List[Dict[str, Any]],
    project_alias_map: Dict[str, str],
    employee_alias_map: Dict[str, str],
) -> Tuple[Any, Optional[int]]:
    for rec in records:
        # Ignore Summary / empty project columns.
        if not rec["project"] or norm_basic(rec["project"]) == "summary":
            continue
        if employee_matches(employee, rec["assignee"], employee_alias_map) and project_matches(project, rec["project"], project_alias_map):
            return rec["value"], rec["col"]
    return "0", None


def update_eis_web(
    wb: openpyxl.Workbook,
    records: List[Dict[str, Any]],
    project_alias_map: Dict[str, str],
    employee_alias_map: Dict[str, str],
) -> int:
    if EIS_WEB_SHEET not in wb.sheetnames:
        # If no EIS_Web sheet, do not create it because user likely has a specific template.
        return 0

    ws = wb[EIS_WEB_SHEET]
    updated = 0

    for row in range(EIS_WEB_START_ROW, ws.max_row + 1):
        employee = cell_text(ws.cell(row, EIS_WEB_EMPLOYEE_COL).value)
        project = cell_text(ws.cell(row, EIS_WEB_PROJECT_COL).value)

        # Skip empty rows.
        if not employee and not project:
            continue
        if not employee or not project:
            ws.cell(row, EIS_WEB_OUTPUT_COL).value = "0"
            updated += 1
            continue

        value, _matched_col = find_eis_web_value(employee, project, records, project_alias_map, employee_alias_map)
        ws.cell(row, EIS_WEB_OUTPUT_COL).value = value
        updated += 1

    return updated


# -----------------------------
# Main process
# -----------------------------

def process(input_file: Path, output_file: Path, eis_sheet: Optional[str] = None) -> None:
    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    # wb_formulas preserves formulas/styles and is used for saving.
    wb = openpyxl.load_workbook(input_file)

    # wb_values reads cached values from formulas. If Excel has recalculated/saved the workbook,
    # row 19 formula results can be read from this workbook.
    wb_values = openpyxl.load_workbook(input_file, data_only=True)

    project_alias_map, project_code_map = load_project_alias_and_codes(wb)
    employee_alias_map = load_alias_sheet(wb, EMPLOYEE_ALIAS_SHEET)

    eis_ws = find_eis_sheet(wb, eis_sheet)
    eis_values_ws = wb_values[eis_ws.title]

    records = build_eis_column_records(eis_ws, eis_values_ws, project_alias_map)

    extra_count = update_eis_extra(wb, records, project_alias_map, project_code_map)
    web_count = update_eis_web(wb, records, project_alias_map, employee_alias_map)

    # Ask Excel to recalculate formulas when opened.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(output_file)

    print("========================================")
    print("EIS Extra/Web Post Process Finished")
    print("========================================")
    print(f"Input : {input_file}")
    print(f"Output: {output_file}")
    print(f"EIS sheet: {eis_ws.title}")
    print(f"EIS_Extra rows updated: {extra_count}")
    print(f"EIS_Web rows updated  : {web_count}")
    print(f"Project aliases loaded : {len(project_alias_map) // 2 if project_alias_map else 0}")
    print(f"Project codes loaded   : {len(project_code_map) // 2 if project_code_map else 0}")
    print(f"Employee aliases loaded: {len(employee_alias_map) // 2 if employee_alias_map else 0}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Post-process EIS_calculated.xlsx and fill EIS_Extra / EIS_Web.")
    parser.add_argument("--input", default=DEFAULT_INPUT_FILE, help="Input reviewed workbook. Default: EIS_calculated.xlsx")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_FILE, help="Output workbook. Default: EIS_final.xlsx")
    parser.add_argument("--eis-sheet", default=None, help="EIS main sheet name. Optional; auto-detect if omitted.")
    args = parser.parse_args()

    process(Path(args.input), Path(args.output), args.eis_sheet)


if __name__ == "__main__":
    main()
